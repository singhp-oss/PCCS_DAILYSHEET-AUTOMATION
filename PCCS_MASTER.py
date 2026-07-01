"""
PCCS MASTER — AWB Automation System
=====================================
Ek hi script mein sab kuch:
  ✈ System Tray icon
  📂 File Watcher (Phase 1)
  📊 Data Extractor (Phase 2)
  📋 Daily Sheet Filler (Phase 3)
  💾 Session Memory
  🔔 Duplicate Alerts
  🆕 CCU + BLR Dual Origin Support

Install: pip install watchdog pypdf openpyxl pywin32 pystray Pillow
Run    : python PCCS_MASTER.py  (ya START.bat)
"""

import os, re, sys, json, time, shutil, glob, threading, queue
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
import difflib
import win32gui
import win32con
from datetime import datetime
from copy import copy

# ─── CONFIG LOAD ──────────────────────────────────────────────────────────────

# EXE ke andar bundled files ka path
if getattr(sys, "_MEIPASS", None):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Config + memory hamesha EXE ke saath wali directory mein
_EXE_DIR    = os.path.dirname(os.path.abspath(sys.argv[0]))
CONFIG_FILE = os.path.join(_EXE_DIR, "config.json")
MEMORY_FILE = os.path.join(_EXE_DIR, "session_memory.json")

def _resource(relative: str) -> str:
    """Bundled file ka sahi path — EXE ya normal script dono ke liye."""
    return os.path.join(BASE_DIR, relative)

def load_config() -> dict:
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

# ── Global config vars — _init_globals() se set honge ────────────────────────
CFG = P = S = None
WATCH_FOLDER = EXTRACT_PATH = PROCESSED_FOLDER = DAILY_PATH = ""
BLR_PROCESSED_FOLDER = ""
EXTRACT_SHEET = DAILY_SHEET_NAME = ""
BLR_EXTRACT_SHEET = BLR_DAILY_SHEET = ""
MACRO_WAIT = 1.0
FLAT_RATE_LIMIT = 1000

def _init_globals():
    """Config.json se sab globals load karo — wizard ke baad bhi call hota hai."""
    global CFG, P, S
    global WATCH_FOLDER, EXTRACT_PATH, PROCESSED_FOLDER, DAILY_PATH
    global BLR_PROCESSED_FOLDER
    global EXTRACT_SHEET, DAILY_SHEET_NAME, MACRO_WAIT, FLAT_RATE_LIMIT
    global BLR_EXTRACT_SHEET, BLR_DAILY_SHEET
    global AGENT_LIST_CCU, AGENT_LIST_BLR, AGENT_LIST

    CFG = load_config()
    P   = CFG["paths"]
    S   = CFG["settings"]

    WATCH_FOLDER         = P["watch_folder"]
    EXTRACT_PATH         = P["extract_excel"]
    PROCESSED_FOLDER     = P["processed_folder"]
    BLR_PROCESSED_FOLDER = P["blr_processed_folder"]
    DAILY_PATH           = P["daily_sheet"]
    EXTRACT_SHEET        = S["extract_sheet"]
    DAILY_SHEET_NAME     = S["daily_sheet_name"]
    MACRO_WAIT           = max(float(S["macro_wait_sec"]), 1.0)
    FLAT_RATE_LIMIT      = int(S["flat_rate_threshold"])
    BLR_EXTRACT_SHEET    = S["blr_extract_sheet"]
    BLR_DAILY_SHEET      = S["blr_sheet_name"]

    AGENT_LIST_CCU = load_agents("CCU")
    AGENT_LIST_BLR = load_agents("BLR")
    AGENT_LIST     = AGENT_LIST_CCU
    print(f"[CONFIG] Loaded — CCU agents:{len(AGENT_LIST_CCU)}  BLR agents:{len(AGENT_LIST_BLR)}")

# ─── AGENTS MASTER ────────────────────────────────────────────────────────────

AGENTS_MASTER_PATH = _resource("AGENTS_MASTER.xlsx")

# ── PATCH 2: load_agents — sheet_name parameter added ─────────────────────────
def load_agents(sheet_name: str = "CCU") -> list:
    """AGENTS_MASTER.xlsx ke specified sheet se agents load karo."""
    agents = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(AGENTS_MASTER_PATH, read_only=True, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            print(f"[AGENTS] Sheet '{sheet_name}' nahi mili — active sheet use kar raha hoon.")
        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[0]
            if val and str(val).strip():
                name = str(val).strip().upper()
                if name not in seen:
                    seen.add(name)
                    agents.append(name)
        wb.close()
        print(f"[AGENTS] {sheet_name}: {len(agents)} agents loaded.")
    except FileNotFoundError:
        print(f"[AGENTS] AGENTS_MASTER.xlsx nahi mili — suggestions off.")
    except Exception as e:
        print(f"[AGENTS] {sheet_name} load error: {e}")
    return sorted(agents)

AGENT_LIST_CCU = load_agents("CCU")
AGENT_LIST_BLR = load_agents("BLR")
AGENT_LIST     = AGENT_LIST_CCU   # backward compat

# ──────────────────────────────────────────────────────────────────────────────

# IndiGo: 312-27706372  |  Star Air: SDG-00007324
AWB_PATTERN = re.compile(r"\b([A-Z0-9]{3}-\d{8})\b")
AWB_SIGNALS = ["interglobe aviation", "air consignment note", "airway bill",
               "ghodawat enterprises", "star air"]

CARGO_PATTERN = re.compile(
    r"incl\. Dimensions or Volume\)\s*\n"
    r"(\d+)\n"                          # group 1 = Pcs
    r"[\d.]+\n"                         # Gross Wt (ignore)
    r"K\n"                               # kg unit
    r"[A-Z]\n"                           # Rate code — single letter (Q/M/C)
    r"([A-Z0-9 ]*)\n"                    # group 2 = Commodity Item (GEN/PER/FSD or blank)
    r"([\d.]+)\n"                       # group 3 = Chargeable Weight
    r"([\d.]+)\n"                       # group 4 = Rate/Charge
    r"[\d,]+\.?\d*\n"                 # Total (ignore)
    r"([A-Z][^\n]{1,60}?)(?:\s*DIMS:|\n)"  # group 5 = Material description
)

# ── CCU Column mappings (1-based) ─────────────────────────────────────────────
# AWB_EXTRACT → CCU DAILYSHEET
STEP1       = [(6, 12)]                                          # F→L  (Agent, macro trigger)
STEP2       = [(7,13),(8,14),(9,15),(10,16),(11,18),(12,19),(13,32)]  # G-L→M-S + M→AF(commodity)
STEP3       = [(1,2),(2,3),(3,5),(4,6)]                         # A-D → B,C,E,F
STEP_LAST   = [(5,11)]                                           # E→K  (Consignor, last)
DAILY_AWB_COL   = 2   # B
DAILY_AGENT_COL = 12  # L
RATE_COL_EXT    = 10  # J in extract
BASIC_COL_DAILY = 17  # Q in daily
DUE_CARR_COL    = 12  # L in extract

# ── PATCH 3: BLR Column mappings ──────────────────────────────────────────────
# AWB_EXTRACT → BLR DAILYSHEET
# Fill order: AWB(B) first → macro → Agent(J)..Rate(N) → Date/Dest/Flight → Consignor(I) last
BLR_STEP1     = [(1, 2)]                                             # A→B  AWB (macro trigger)
BLR_STEP2     = [(6,10),(7,11),(8,12),(9,13),(10,14),(11,16),(12,17),(13,33)]  # +commodity→AG(33)
BLR_STEP3     = [(2,3),(3,5),(4,6)]                                  # Date→C, Dest→E, Flight→F
BLR_STEP_LAST = [(5,9)]                                              # E→I  Consignor (last)
BLR_DAILY_AWB_COL   = 2   # B
BLR_DAILY_AGENT_COL = 10  # J
BLR_RATE_COL_EXT    = 10  # J in extract
BLR_BASIC_COL_DAILY = 15  # O in daily (flat rate)

# Queue: watcher → filler thread
_fill_queue  = queue.Queue()
_seen        = set()
_seen_lock   = threading.Lock()

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _fatal(msg: str):
    try:
        root = tk.Tk(); root.withdraw()
        messagebox.showerror("PCCS Error", msg)
        root.destroy()
    except Exception:
        print(f"[FATAL] {msg}")
    sys.exit(1)

def _popup(kind: str, title: str, msg: str, buttons=None):
    result = [None]
    ev = threading.Event()
    def _show():
        root = tk.Tk(); root.withdraw()
        root.attributes("-topmost", True)
        if kind == "error":
            messagebox.showerror(title, msg, parent=root)
        elif kind == "warn":
            messagebox.showwarning(title, msg, parent=root)
        elif kind == "yesno":
            result[0] = messagebox.askyesno(title, msg, parent=root)
        else:
            messagebox.showinfo(title, msg, parent=root)
        root.destroy()
        ev.set()
    try:
        if threading.current_thread() is threading.main_thread():
            _show()
        else:
            _popup_queue.put(_show)
            ev.wait(timeout=60)
    except Exception:
        pass
    return result[0]

_popup_queue = queue.Queue()

# ─── SESSION MEMORY ───────────────────────────────────────────────────────────

def memory_load() -> dict:
    if os.path.isfile(MEMORY_FILE):
        try:
            with open(MEMORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"last_awb": "", "last_agent": "", "last_time": "", "processed_awbs": []}

_memory_lock = threading.Lock()

def memory_save(data: dict):
    tmp = MEMORY_FILE + ".tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        os.replace(tmp, MEMORY_FILE)
    except Exception as e:
        print(f"[MEMORY] Save failed: {e}")
        try:
            if os.path.exists(tmp): os.remove(tmp)
        except Exception:
            pass

def memory_add_awb(awb_no: str, agent: str):
    with _memory_lock:
        mem = memory_load()
        mem["last_awb"]   = awb_no
        mem["last_agent"] = agent
        mem["last_time"]  = datetime.now().strftime("%d-%m-%Y %H:%M")
        if awb_no not in mem["processed_awbs"]:
            mem["processed_awbs"].append(awb_no)
        memory_save(mem)

def memory_show_last():
    mem = memory_load()
    if mem["last_awb"]:
        msg = (f"Last processed AWB:\n\n"
               f"  AWB No  : {mem['last_awb']}\n"
               f"  Agent   : {mem['last_agent']}\n"
               f"  Time    : {mem['last_time']}\n\n"
               f"Total processed: {len(mem['processed_awbs'])} AWBs")
    else:
        msg = "Abhi tak koi AWB process nahi hua."
    _popup("info", "Last Session Info", msg)

def memory_clear_processed():
    mem   = memory_load()
    count = len(mem.get("processed_awbs", []))
    ans = _popup("yesno", "🗑 Memory Reset",
                 f"Session memory mein {count} processed AWB(s) hain.\n\n"
                 "Reset karne se:\n"
                 "  • Duplicate AWB alerts band ho jayenge\n"
                 "  • Same AWBs dobara process ho sakte hain\n\n"
                 "Reset karna chahte ho?")
    if ans:
        with _memory_lock:
            mem = memory_load()
            mem["processed_awbs"] = []
            memory_save(mem)
        _popup("info", "Reset ✔",
               f"{count} AWB(s) memory se clear ho gaye.\n"
               "Ab same AWBs dobara process ho sakte hain.")

# ─── PDF DETECTION ────────────────────────────────────────────────────────────

def _wait_file(path: str, timeout: float = 30.0) -> bool:
    """
    File download complete hone tak wait karo.
    Size 3 baar same aaye = download done.
    Slow internet ke liye 30 sec timeout.
    """
    deadline    = time.time() + timeout
    last_size   = -1
    stable_hits = 0
    while time.time() < deadline:
        try:
            if os.path.exists(path):
                sz = os.path.getsize(path)
                if sz > 100:
                    if sz == last_size:
                        stable_hits += 1
                        if stable_hits >= 3:   # 3 × 0.5s = 1.5s stable = done
                            return True
                    else:
                        stable_hits = 0
                    last_size = sz
        except OSError:
            pass
        time.sleep(0.5)
    return last_size > 100   # timeout par bhi file hai to try karo

def read_pdf_page1(path: str) -> str:
    try:
        from pypdf import PdfReader
        return PdfReader(path).pages[0].extract_text() or ""
    except Exception:
        return ""

def is_awb_pdf(path: str):
    """(awb_no, text) ya (None, None)"""
    if not _wait_file(path):
        return None, None
    text = ""
    for _ in range(15):          # 15 × 0.5s = 7.5s max retry
        text = read_pdf_page1(path)
        if text.strip():
            break
        time.sleep(0.5)
    if not text.strip():
        return None, None
    tl = text.lower()
    if not any(s in tl for s in AWB_SIGNALS):
        return None, None
    m = AWB_PATTERN.search(text)
    return (m.group(1), text) if m else (None, None)

def parse_extras(text: str):
    dest = date = ""
    m = re.search(r"Airport of Destination\s*\n([A-Z]{3})-", text)
    if m:
        dest = m.group(1)
    else:
        m = re.search(r"\n\s*([A-Z]{3})\s*\n6E\b", text)
        if m: dest = m.group(1)
    m = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", text)
    if m: date = m.group(1)
    return dest, date

# ── PATCH 4: detect_origin — nayi function ────────────────────────────────────
def detect_origin(text: str) -> str:
    """
    PDF text se origin airport detect karo.
    Method 1: PDF header  "312|BLR|27706372"  ya  "SDG|CCU|00007324"
    Method 2: Airport of Departure field  "BLR-BENGALURU"
    Returns: "CCU" / "BLR" / "" (unknown)
    """
    # Method 1 — PDF top line: IndiGo "312|CCU|..." ya Star Air "SDG|CCU|..."
    m = re.search(r"[A-Z0-9]{3}\|([A-Z]{3})\|[A-Z0-9]+", text)
    if m:
        return m.group(1).upper()
    # Method 2 — Airport of Departure field
    m = re.search(r"Airport of Departure[^\n]*\n([A-Z]{3})-", text)
    if m:
        return m.group(1).upper()
    return ""

# ── PATCH 5: get_origin_cfg — origin ke hisaab se config ─────────────────────
def get_origin_cfg(origin: str) -> dict:
    """
    Origin string ke basis par sahi config dict return karo.
    Default: CCU (agar unknown origin aaye).
    """
    if origin == "BLR":
        return {
            "label"           : "BLR",
            "extract_sheet"   : BLR_EXTRACT_SHEET,
            "processed_folder": BLR_PROCESSED_FOLDER,
            "daily_sheet_name": BLR_DAILY_SHEET,
            "agent_list"      : AGENT_LIST_BLR,
        }
    return {   # CCU default
        "label"           : "CCU",
        "extract_sheet"   : EXTRACT_SHEET,
        "processed_folder": PROCESSED_FOLDER,
        "daily_sheet_name": DAILY_SHEET_NAME,
        "agent_list"      : AGENT_LIST_CCU,
    }

# ─── RENAME DIALOG ────────────────────────────────────────────────────────────

class RenameDialog:
    BG="0F1117"; CARD="1C1F2E"; BORDER="2A2D45"
    ACCENT="3D7EFF"; ACC_HV="5C96FF"; GREEN="2ECC87"
    TEXT="E8EBF5"; MUTED="737AA8"; EBG="141622"; ERR_BG="2D1A1A"

    # ── PATCH 6: __init__ — agent_list + origin params added ──────────────────
    def __init__(self, awb_no, dest, date, agent_list=None, origin="CCU"):
        self.awb_no     = awb_no
        self.dest       = dest
        self.date       = date
        self.origin     = origin                                    # NEW
        self.agent_list = agent_list if agent_list is not None else AGENT_LIST_CCU  # NEW
        self.agent_name = None
        self._err_job   = None
        self._ac_win    = None
        self._ac_list   = None

    def show(self):
        self._build()
        self.root.mainloop()
        return self.agent_name

    def _build(self):
        self.root = tk.Tk()
        self.root.title("AWB Detected")
        self.root.configure(bg=f"#{self.BG}")
        self.root.resizable(False, False)
        self.root.attributes("-topmost", True)
        W,H=430,310
        sw=self.root.winfo_screenwidth(); sh=self.root.winfo_screenheight()
        self.root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2-40}")

        tk.Frame(self.root, bg=f"#{self.ACCENT}", height=3).pack(fill="x")
        card = tk.Frame(self.root, bg=f"#{self.CARD}", padx=26, pady=20)
        card.pack(fill="both", expand=True, padx=10, pady=(8,10))

        hdr = tk.Frame(card, bg=f"#{self.CARD}")
        hdr.pack(fill="x")
        iw = tk.Frame(hdr, bg="#182840", width=44, height=44)
        iw.pack(side="left", padx=(0,12)); iw.pack_propagate(False)
        tk.Label(iw, text="✈", font=("Segoe UI Emoji",19),
                 bg="#182840", fg=f"#{self.ACCENT}").place(relx=.5,rely=.5,anchor="center")
        rt = tk.Frame(hdr, bg=f"#{self.CARD}"); rt.pack(side="left")
        tk.Label(rt, text="Air Waybill Detected",
                 font=("Segoe UI",13,"bold"),
                 bg=f"#{self.CARD}", fg=f"#{self.TEXT}").pack(anchor="w")
        # ── PATCH 7: Subtitle mein origin label add kiya ──────────────────────
        tk.Label(rt, text=f"IndiGo Cargo  •  PCCS  •  {self.origin}",
                 font=("Segoe UI",8),
                 bg=f"#{self.CARD}", fg=f"#{self.MUTED}").pack(anchor="w")

        tk.Frame(card, bg=f"#{self.BORDER}", height=1).pack(fill="x", pady=(12,10))

        badge = tk.Frame(card, bg=f"#{self.EBG}", pady=8, padx=14)
        badge.pack(fill="x")
        tk.Label(badge, text="AWB NUMBER", font=("Segoe UI",7,"bold"),
                 bg=f"#{self.EBG}", fg=f"#{self.MUTED}").pack(anchor="w")
        tk.Label(badge, text=self.awb_no, font=("Consolas",20,"bold"),
                 bg=f"#{self.EBG}", fg=f"#{self.GREEN}").pack(anchor="w")
        if self.dest or self.date:
            row = tk.Frame(badge, bg=f"#{self.EBG}"); row.pack(anchor="w", pady=(3,0))
            if self.dest:
                tk.Label(row, text=f"TO: {self.dest}", font=("Segoe UI",8),
                         bg=f"#{self.EBG}", fg=f"#{self.MUTED}").pack(side="left", padx=(0,14))
            if self.date:
                tk.Label(row, text=f"DATE: {self.date}", font=("Segoe UI",8),
                         bg=f"#{self.EBG}", fg=f"#{self.MUTED}").pack(side="left")

        tk.Label(card, text="Agent Name", font=("Segoe UI",9),
                 bg=f"#{self.CARD}", fg=f"#{self.MUTED}").pack(anchor="w", pady=(12,3))

        self._ef = tk.Frame(card, bg=f"#{self.ACCENT}", padx=1, pady=1)
        self._ef.pack(fill="x")
        self._var = tk.StringVar()
        self._entry = tk.Entry(self._ef, textvariable=self._var,
                               font=("Segoe UI", 13, "bold"),
                               bg=f"#{self.EBG}", fg=f"#{self.TEXT}",
                               insertbackground=f"#{self.ACCENT}",
                               bd=0, highlightthickness=0, relief="flat")
        self._entry.pack(fill="x", ipady=8, padx=10)
        self._entry.bind("<KeyRelease>", self._on_keyrelease)
        self._entry.bind("<Down>",       self._ac_focus_list)
        self._entry.bind("<Return>",     lambda _: self._submit())
        self._entry.bind("<Escape>",     self._on_escape)
        self._entry.bind("<FocusIn>",    lambda _: self._reset_border())

        self._hint_var = tk.StringVar(value="")
        tk.Label(card, textvariable=self._hint_var,
                 bg=f"#{self.CARD}", fg="#FF7070",
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(2, 0))

        self._ac_win  = None
        self._ac_list = None

        last = memory_load().get("last_agent", "")
        if last:
            self._var.set(last.upper())
            self.root.after(10, lambda: (
                self._entry.selection_range(0, "end"),
                self._entry.icursor("end")))

        btns = tk.Frame(card, bg=f"#{self.CARD}"); btns.pack(fill="x", pady=(12,0))
        tk.Button(btns, text="Cancel", font=("Segoe UI",9),
                  bg=f"#{self.CARD}", fg=f"#{self.MUTED}",
                  activebackground=f"#{self.CARD}", activeforeground=f"#{self.TEXT}",
                  bd=0, padx=14, pady=7, cursor="hand2",
                  command=self._cancel).pack(side="right", padx=(6,0))
        sb = tk.Button(btns, text="  Save File  →",
                       font=("Segoe UI",10,"bold"),
                       bg=f"#{self.ACCENT}", fg="white",
                       activebackground=f"#{self.ACC_HV}", activeforeground="white",
                       bd=0, padx=20, pady=7, cursor="hand2", command=self._submit)
        sb.pack(side="right")
        sb.bind("<Enter>", lambda _: sb.config(bg=f"#{self.ACC_HV}"))
        sb.bind("<Leave>", lambda _: sb.config(bg=f"#{self.ACCENT}"))

        self.root.after(300, self._force_focus)

    def _force_focus(self):
        import ctypes
        try:
            hwnd = self.root.winfo_id()
            ctypes.windll.user32.keybd_event(0x12, 0, 0, 0)
            ctypes.windll.user32.keybd_event(0x12, 0, 0x0002, 0)
            ctypes.windll.user32.ShowWindow(hwnd, 9)
            ctypes.windll.user32.SetForegroundWindow(hwnd)
        except Exception: pass
        try:
            self.root.lift()
            self.root.attributes("-topmost", True)
            self.root.focus_force()
            self._entry.focus_force()
            if self._var.get():
                self._entry.selection_range(0, "end")
                self._entry.icursor("end")
            else:
                self._entry.icursor("end")
        except Exception: pass
        self.root.after(500, self._retry_focus)

    def _retry_focus(self):
        import ctypes
        try:
            hwnd = self.root.winfo_id()
            ctypes.windll.user32.SetForegroundWindow(hwnd)
            self._entry.focus_force()
        except Exception: pass

    def _open_dropdown(self, event=None): pass
    def _on_selected(self, event=None):   pass

    # ── PATCH 8: _on_keyrelease — self.agent_list use karo ────────────────────
    def _on_keyrelease(self, event):
        if event.keysym in ("Down", "Up", "Return", "Escape", "Tab"):
            return
        try:
            cur = self._entry.index(tk.INSERT)
        except Exception:
            cur = "end"
        typed = self._var.get().upper()
        if self._var.get() != typed:
            self._var.set(typed)
            try: self._entry.icursor(cur)
            except Exception: pass

        if not typed:
            self._ac_hide()
            self._hint_var.set("")
            return

        # self.agent_list use karo (CCU ya BLR ke hisaab se)
        filtered = [a for a in self.agent_list if typed in a]
        if filtered:
            self._hint_var.set(f"↓ {len(filtered)} match — Down arrow ya click")
            self._ac_show(filtered)
        else:
            self._ac_hide()
            close = difflib.get_close_matches(typed, self.agent_list, n=1, cutoff=0.55)
            if close:
                self._hint_var.set(f"Did you mean: {close[0]}?")
                self._ac_show(close)
            else:
                self._hint_var.set("No match found")

    # ── PATCH 9: _ac_show — self.agent_list use karo ──────────────────────────
    def _ac_show(self, matches):
        self._ac_hide()
        if not matches or not self.agent_list:   # self.agent_list check
            return
        self.root.update_idletasks()
        try:
            ex = self._ef.winfo_rootx()
            ey = self._ef.winfo_rooty() + self._ef.winfo_height()
            ew = self._ef.winfo_width()
        except Exception:
            return

        items  = matches[:9]
        row_h  = 26
        border = 2
        h      = len(items) * row_h + border * 2

        self._ac_win = tk.Toplevel(self.root)
        self._ac_win.overrideredirect(True)
        self._ac_win.attributes("-topmost", True)
        self._ac_win.geometry(f"{ew}x{h}+{ex}+{ey}")
        self._ac_win.configure(bg=f"#{self.ACCENT}")

        self._ac_list = tk.Listbox(
            self._ac_win,
            font=("Segoe UI", 11, "bold"),
            bg=f"#{self.EBG}",
            fg=f"#{self.TEXT}",
            selectbackground=f"#{self.ACCENT}",
            selectforeground="white",
            activestyle="none",
            borderwidth=0,
            highlightthickness=0,
            relief="flat",
            height=len(items)
        )
        self._ac_list.pack(fill="both", expand=True, padx=1, pady=1)

        for m in items:
            self._ac_list.insert("end", f"  {m}")

        self._ac_list.bind("<Return>",          self._ac_select)
        self._ac_list.bind("<ButtonRelease-1>",  self._ac_select)
        self._ac_list.bind("<Escape>",           lambda e: self._ac_hide_focus())
        self._ac_list.bind("<Up>",               self._ac_up)

    def _ac_hide(self):
        try:
            if self._ac_win:
                self._ac_win.destroy()
                self._ac_win  = None
                self._ac_list = None
        except Exception:
            pass

    def _ac_hide_focus(self):
        self._ac_hide()
        try: self._entry.focus_set()
        except Exception: pass

    def _ac_focus_list(self, event=None):
        if self._ac_list:
            self._ac_list.focus_set()
            self._ac_list.selection_clear(0, "end")
            self._ac_list.selection_set(0)
            self._ac_list.activate(0)
            return "break"

    def _ac_up(self, event=None):
        try:
            idx = self._ac_list.curselection()
            if not idx or idx[0] == 0:
                self._entry.focus_set()
                return "break"
        except Exception:
            pass

    def _ac_select(self, event=None):
        try:
            if self._ac_list:
                sel = self._ac_list.curselection()
                if not sel:
                    sel = (self._ac_list.nearest(event.y),) if event else None
                if sel:
                    val = self._ac_list.get(sel[0]).strip()
                    self._var.set(val)
                    self._ac_hide()
                    self._hint_var.set(f"✔ {val}")
                    self.root.after(60, self._submit)
        except Exception:
            pass
        return "break"

    def _on_escape(self, event=None):
        if self._ac_win:
            self._ac_hide_focus()
        else:
            self._cancel()

    def _submit(self):
        name = self._var.get().strip().upper()
        if not name:
            self._ef.config(bg="#FF4444")
            self._hint_var.set("⚠ Agent naam daalna zaroori hai!")
            self._entry.focus_set()
            if self._err_job: self.root.after_cancel(self._err_job)
            self._err_job = self.root.after(700, self._reset_border)
            return
        self.agent_name = name; self.root.destroy()

    def _cancel(self):
        self._ac_hide()
        self.root.destroy()

    def _reset_border(self):
        self._ef.config(bg=f"#{self.ACCENT}")
        self._hint_var.set("")

# ─── NOTIFICATION SYSTEM — tkinter-free ──────────────────────────────────────
# Background threads mein tk.Tk() banana dangerous hai — pystray notify use karo

_tray_icon_ref = None   # build_tray() se set hoga

def _notify(title: str, msg: str):
    """
    Pystray system tray notification — tkinter se bilkul alag.
    Thread-safe. Background threads se safely call ho sakta hai.
    """
    try:
        if _tray_icon_ref:
            _tray_icon_ref.notify(msg, title)
        else:
            print(f"[NOTIFY] {title}: {msg}")
    except Exception as e:
        print(f"[NOTIFY] {title}: {msg}  (err: {e})")

def show_toast(filename: str, origin: str = ""):
    """Renamed file ka notification."""
    label = f"File Saved [{origin}]" if origin else "File Saved"
    _notify(f"✔  {label}", filename)

# ─── PHASE 1: WATCHER ─────────────────────────────────────────────────────────

def do_rename(original: str, agent: str, awb_no: str) -> str:
    folder = os.path.dirname(original)
    safe   = re.sub(r'[\\/*?:"<>|\s]+', "_", agent.strip().upper()).strip("_")
    new_name = f"{safe}_{awb_no}.pdf"
    new_path = os.path.join(folder, new_name)
    i = 1
    while os.path.exists(new_path):
        new_name = f"{safe}_{awb_no}_{i}.pdf"
        new_path = os.path.join(folder, new_name)
        i += 1
    os.rename(original, new_path)
    return new_path

# IndiGo: AGENT_312-27706372.pdf  |  Star Air: AGENT_SDG-00007324.pdf
ALREADY_RENAMED = re.compile(r"^.+_[A-Z0-9]{3}-\d{8}\.pdf$", re.IGNORECASE)

_rename_lock = threading.Lock()

# ── PATCH 11: _show_rename_dialog — origin param added ────────────────────────
def _show_rename_dialog(awb_no: str, dest: str, date: str, origin: str = "CCU"):
    """
    RenameDialog main thread mein chalao.
    Origin ke hisaab se correct agent_list pass karo.
    """
    cfg    = get_origin_cfg(origin)
    result = [None]
    ev     = threading.Event()
    def _show():
        try:
            d         = RenameDialog(awb_no, dest, date,
                                     agent_list=cfg["agent_list"],
                                     origin=origin)
            result[0] = d.show()
        except Exception as e:
            print(f"[DIALOG] Error: {e}")
        finally:
            ev.set()
    _popup_queue.put(_show)
    ev.wait(timeout=300)
    return result[0]

# ── PATCH 12: _ask_origin_main — unknown origin ke liye popup ─────────────────
def _ask_origin_main(awb_no: str) -> str:
    """Origin detect nahi hua — user se pucho."""
    result = [None]
    ev = threading.Event()
    def _show():
        root = tk.Tk(); root.withdraw()
        root.attributes("-topmost", True)
        ans = messagebox.askyesno(
            "⚠ Origin Select Karo",
            f"AWB: {awb_no}\n\n"
            "Origin automatically detect nahi hua.\n\n"
            "CCU (Kolkata) hai?\n"
            "Yes = CCU   |   No = BLR",
            parent=root
        )
        result[0] = "CCU" if ans else "BLR"
        root.destroy(); ev.set()
    _popup_queue.put(_show)
    ev.wait(timeout=30)
    return result[0] or "CCU"

def _ask_yesno_main(title, msg):
    result = [None]
    ev = threading.Event()
    def _show():
        root = tk.Tk(); root.withdraw()
        root.attributes("-topmost", True)
        result[0] = messagebox.askyesno(title, msg, parent=root)
        root.destroy(); ev.set()
    _popup_queue.put(_show)
    ev.wait(timeout=30)
    return result[0]

# ── PATCH 13: handle_new_pdf — origin routing added ───────────────────────────
def handle_new_pdf(path: str):
    """Watcher thread → PDF detect → origin check → dialog → rename."""
    try:
        filename = os.path.basename(path)
        print(f"\n[PDF] {filename}")

        if ALREADY_RENAMED.match(filename):
            print(f"[AUTO] Already renamed — popup skip, seedha process.")
            agent    = filename.rsplit("_", 1)[0].replace("_", " ").strip()
            # IndiGo: 312-27706372  |  Star Air: SDG-00007324
            awb_full  = re.search(r"([A-Z0-9]{3}-\d{8})", filename)
            short_awb = awb_full.group(1).split("-")[1] if awb_full else ""
            with _seen_lock:
                _seen.add(path)
            memory_add_awb(short_awb, agent)
            show_toast(filename)
            return

        awb_no, text = is_awb_pdf(path)
        if not awb_no:
            print("[SKIP] AWB nahi — on_modified retry ke liye _seen se hata raha hoon.")
            with _seen_lock:
                _seen.discard(path)   # ← on_modified dobara try kar sakta hai
            return

        # ── Origin detect ──────────────────────────────────────────────────────
        origin = detect_origin(text) if text else ""
        if origin not in ("CCU", "BLR"):
            print(f"[ORIGIN] Unknown '{origin}' — user se puch raha hoon.")
            origin = _ask_origin_main(awb_no)
        print(f"[ORIGIN] {origin}")

        # Duplicate check
        mem = memory_load()
        short_awb = awb_no.split("-")[1] if "-" in awb_no else awb_no
        if short_awb in mem.get("processed_awbs", []):
            ans = _ask_yesno_main(
                "⚠ Duplicate AWB",
                f"AWB {short_awb} pehle bhi process hua hai!\n\n"
                f"Phir bhi rename karna chahte ho?"
            )
            if not ans:
                print(f"[SKIP] Duplicate AWB {short_awb} — skip.")
                return

        dest, date = parse_extras(text)
        print(f"[AWB] {awb_no}  DEST:{dest}  DATE:{date}  ORIGIN:{origin}")

        with _rename_lock:
            agent = _show_rename_dialog(awb_no, dest, date, origin=origin)
        if not agent:
            print("[SKIP] Cancel.")
            return

        if not os.path.exists(path):
            print(f"[ERR] File nahi mili: {path}")
            return

        new_path = do_rename(path, agent, awb_no)
        new_name = os.path.basename(new_path)
        with _seen_lock:
            _seen.add(new_path)

        print(f"[SAVED] → {new_name}  |  Queue mein daala...")
        _start_auto_fill_worker()
        _auto_fill_queue.put((new_path, origin, agent))

    except Exception as e:
        print(f"[ERR] handle_new_pdf: {e}")
    finally:
        with _seen_lock:
            _seen.discard(path)

from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

class PDFHandler(FileSystemEventHandler):
    # Debounce — on_modified spam rok
    _pending: dict = {}
    _pending_lock  = threading.Lock()

    def on_created(self, event):
        if not event.is_directory: self._try(event.src_path)
    def on_moved(self, event):
        if not event.is_directory: self._try(event.dest_path)
    def on_modified(self, event):
        """
        Download complete hone par modify event aata hai.
        Debounce: 2 sec baad process karo — spam events ignore honge.
        """
        if event.is_directory: return
        path = event.src_path
        if not path.lower().endswith(".pdf"): return
        with self._pending_lock:
            # Pehle se pending timer hai to cancel karo
            old = self._pending.get(path)
            if old:
                old.cancel()
            # 2 sec baad try karo
            t = threading.Timer(2.0, self._try, args=(path,))
            self._pending[path] = t
            t.start()

    def _try(self, path):
        if not path.lower().endswith(".pdf"): return
        # Cleanup pending dict
        with self._pending_lock:
            self._pending.pop(path, None)
        with _seen_lock:
            if path in _seen: return
            _seen.add(path)
        threading.Thread(target=handle_new_pdf, args=(path,),
                         daemon=True, name=f"PDF-{os.path.basename(path)[:20]}").start()

# ─── AUTO FILL AFTER RENAME ───────────────────────────────────────────────────

_auto_fill_queue  = queue.Queue()   # sequential fill queue
_auto_fill_worker_started = False

def _start_auto_fill_worker():
    """Ek worker thread — queue se ek-ek fill karo."""
    global _auto_fill_worker_started
    if _auto_fill_worker_started:
        return
    _auto_fill_worker_started = True
    def _worker():
        while True:
            try:
                task = _auto_fill_queue.get(timeout=2)
                if task is None:
                    break
                pdf_path, origin, agent = task
                try:
                    auto_fill_single(pdf_path, origin, agent)
                except Exception as e:
                    print(f"[AUTO-FILL WORKER] {e}")
                    _notify("❌ Auto Fill Error", str(e)[:80])
            except queue.Empty:
                pass
    threading.Thread(target=_worker, daemon=True, name="AutoFillWorker").start()
    print("[AUTO-FILL] Worker thread started")

def auto_fill_single(pdf_path: str, origin: str, agent: str):
    """Queue worker mein chalta hai — sequential, thread-safe."""
    import pythoncom
    pythoncom.CoInitialize()
    awb_display = ""
    excel       = None
    opened_new  = False
    try:
        print(f"[AUTO-FILL] Extracting: {os.path.basename(pdf_path)}")
        data = extract_from_pdf(pdf_path)
        if not data or not data.get("awb_no"):
            _notify("❌ Extract Fail",
                    f"{os.path.basename(pdf_path)} — manual fill karo")
            return

        awb_display = data["awb_no"]
        cfg         = get_origin_cfg(origin)

        result = save_to_extract(data, cfg["extract_sheet"])
        if result == "error":
            _notify("❌ Extract Save Fail", f"AWB {awb_display}")
            return

        import win32com.client
        daily_name = os.path.basename(DAILY_PATH).lower()
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            opened_new    = True

        excel.DisplayAlerts = False
        excel.EnableEvents  = True

        wb_daily = None
        for wb in excel.Workbooks:
            if os.path.basename(wb.FullName).lower() == daily_name:
                wb_daily = wb; break
        if wb_daily is None:
            wb_daily = excel.Workbooks.Open(DAILY_PATH)

        sheet_name = cfg["daily_sheet_name"]
        ws_daily   = None
        for sh in wb_daily.Sheets:
            if sh.Name.strip().lower() == sheet_name.strip().lower():
                ws_daily = sh; break

        if ws_daily is None:
            _notify("❌ Sheet Nahi Mila",
                    f"AWB {awb_display} — '{sheet_name}'")
            return

        existing = _daily_existing_awbs(ws_daily)
        if awb_display in existing:
            move_to_processed(pdf_path, cfg["processed_folder"])
            return

        ws_daily.Activate()
        target_row = _daily_next_empty(ws_daily)

        rd = {
            1 : data.get("awb_no",""),      2 : data.get("date",""),
            3 : data.get("dest",""),         4 : data.get("flight",""),
            5 : data.get("consignor",""),    6 : data.get("agent","") or agent,
            7 : data.get("material",""),     8 : data.get("pcs",""),
            9 : data.get("chargeable_wt",""),10: data.get("rate",""),
            11: data.get("due_agent",""),    12: data.get("due_carrier",""),
            13: data.get("commodity_item",""),  # col M → AF(CCU) / AG(BLR)
        }

        fill_fn = fill_daily_row_blr if origin == "BLR" else fill_daily_row
        fill_fn(ws_daily, target_row, rd)
        wb_daily.Save()
        print(f"[AUTO-FILL] ✅ [{origin}] AWB {awb_display} → Row {target_row}")

        move_to_processed(pdf_path, cfg["processed_folder"])
        memory_add_awb(awb_display, agent)
        _notify(f"✔ Filed [{origin}]",
                f"AWB {awb_display}  →  Row {target_row}   •   {agent}")

    except Exception as e:
        print(f"[AUTO-FILL ERROR] {e}")
        _notify("❌ Auto Fill Error",
                f"AWB {awb_display or '?'} — {str(e)[:60]}")
    finally:
        try:
            if excel: excel.DisplayAlerts = True
            if opened_new and excel:
                try:
                    if excel.Workbooks.Count == 0: excel.Quit()
                except Exception: pass
        except Exception: pass
        try: pythoncom.CoUninitialize()
        except Exception: pass

# ─── PHASE 2: EXTRACT ─────────────────────────────────────────────────────────

def extract_from_pdf(pdf_path: str) -> dict:
    filename = os.path.basename(pdf_path)
    agent = filename.rsplit("_",1)[0].replace("_"," ").strip() if "_" in filename else ""
    text  = read_pdf_page1(pdf_path)
    if not text: return {}

    def f(pattern, grp=1):
        m = re.search(pattern, text)
        return m.group(grp).strip() if m else ""

    # IndiGo: 312-27706372  |  Star Air: SDG-00007324
    awb_full    = f(r"\b([A-Z0-9]{3}-\d{8})\b")
    awb_no      = awb_full.split("-")[1] if "-" in awb_full else awb_full
    date_raw    = f(r"\b(\d{2}/\d{2}/\d{4})\b")
    date_fmt    = date_raw.replace("/","-") if date_raw else ""
    dest        = f(r"Airport of Destination\s*\n([A-Z]{3})-") or f(r"\n\s*([A-Z]{3})\s*\n6E\b")
    # Flight — IndiGo: 6E975  |  Star Air: S5620  → Airport of Destination line se nikalo
    m_fl = re.search(r"[A-Z]{3}-[A-Z]+\s+([A-Z0-9]{2,3}\d{3,4})\s+\d{2}/\d{2}/\d{4}", text)
    flight = m_fl.group(1) if m_fl else f(r"\b(6E\d{3,4})\b")
    consignor   = f(r"Shipper.s Name and Address\s*\n([A-Z].+?)\n")
    cargo          = CARGO_PATTERN.search(text)
    # Commodity Item: from cargo table OR SHC: tag in handling info
    commodity_item = (cargo.group(2).strip() if cargo and cargo.group(2) else "")
    if not commodity_item:
        m_shc = re.search(r"SHC:([A-Z]{3}):", text)
        if m_shc:
            commodity_item = m_shc.group(1)   # SHC:GEN: → "GEN"
    pcs            = cargo.group(1) if cargo else ""
    chw            = cargo.group(3) if cargo else ""
    rate           = cargo.group(4) if cargo else ""
    material = ""
    if cargo and cargo.group(5):
        material = cargo.group(5).strip()
        if "DIMS" in material:
            material = material.split("DIMS")[0].strip()
    if not material:
        m2 = re.search(r"\n([A-Z][A-Z0-9 &/(),'.:+-]{2,}?)\s*DIMS:", text)
        if m2:
            material = m2.group(1).strip()
    due_agent = due_carrier_raw = ""
    pradeep_idx = text.rfind("PRADEEP CARGO")
    if pradeep_idx > 0:
        nums = re.findall(r"([\d,]+\.\d+)", text[:pradeep_idx])
        if len(nums) >= 2:
            due_agent       = nums[-2].replace(",","")
            due_carrier_raw = nums[-1].replace(",","")
    if not due_agent:
        due_agent = f(r"DUE AGENT\s*:\s*([\d,]+\.?\d*)").replace(",","")
    if not due_carrier_raw:
        due_carrier_raw = f(r"Total other Charges Due Carrier\s*\n([\d,]+\.?\d*)").replace(",","")

    try:
        due_carrier = str(int(float(due_carrier_raw)))
    except (ValueError, TypeError):
        due_carrier = due_carrier_raw

    return {"awb_no":awb_no,"date":date_fmt,"dest":dest,"flight":flight,
            "consignor":consignor,"agent":agent,"material":material,
            "pcs":pcs,"chargeable_wt":chw,"rate":rate,
            "due_agent":due_agent,"due_carrier":due_carrier,
            "commodity_item":commodity_item}

def _find_last_data_row(ws) -> int:
    last = 1
    for r in range(2, ws.max_row+1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            last = r
    return last

def _copy_row_style(ws, from_row, to_row):
    for col in range(1,14):   # 13 columns: A-M
        src  = ws.cell(row=from_row, column=col)
        dest = ws.cell(row=to_row,   column=col)
        if src.has_style:
            dest.font          = copy(src.font)
            dest.fill          = copy(src.fill)
            dest.border        = copy(src.border)
            dest.alignment     = copy(src.alignment)
            dest.number_format = src.number_format
    ws.row_dimensions[to_row].height = ws.row_dimensions[from_row].height or 18.75

# ── PATCH 14: save_to_extract — sheet_name param added ────────────────────────
def save_to_extract(data: dict, sheet_name: str = None) -> str:
    """'added'/'duplicate'/'error' — sheet_name: CCU ya BLR"""
    if sheet_name is None:
        sheet_name = EXTRACT_SHEET   # default CCU
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXTRACT_PATH)
        ws = wb[sheet_name]

        existing = set()
        for r in range(2, ws.max_row+1):
            v = ws.cell(row=r, column=1).value
            if v: existing.add(str(v).strip())

        if data["awb_no"] in existing:
            wb.close()
            return "duplicate"

        last_row = _find_last_data_row(ws)
        new_row  = last_row + 1

        for i, key in enumerate(["consignor","agent","material","pcs","chargeable_wt","rate"], 5):
            ws.cell(row=new_row, column=i).value = data.get(key,"")
        for i, key in enumerate(["awb_no","date","dest","flight"], 1):
            ws.cell(row=new_row, column=i).value = data.get(key,"")
        ws.cell(row=new_row, column=11).value = data.get("due_agent","")
        ws.cell(row=new_row, column=12).value = data.get("due_carrier","")
        ws.cell(row=new_row, column=13).value = data.get("commodity_item","")  # col M

        for col_i, key in [(8,"pcs"),(9,"chargeable_wt"),(10,"rate"),(11,"due_agent"),(12,"due_carrier")]:
            cell = ws.cell(row=new_row, column=col_i)
            try:
                cell.value = float(cell.value or 0)
                cell.number_format = "#,##0.00"
            except (ValueError, TypeError): pass

        _copy_row_style(ws, 2, new_row)
        wb.save(EXTRACT_PATH)
        return "added"
    except Exception as e:
        print(f"  [EXTRACT ERROR] {e}")
        return "error"

# ── PATCH 15: move_to_processed — folder param added ──────────────────────────
def move_to_processed(pdf_path: str, processed_folder: str = None):
    if processed_folder is None:
        processed_folder = PROCESSED_FOLDER
    os.makedirs(processed_folder, exist_ok=True)
    fname     = os.path.basename(pdf_path)
    dest_path = os.path.join(processed_folder, fname)
    if os.path.exists(dest_path):
        name,ext = os.path.splitext(fname)
        dest_path = os.path.join(processed_folder, f"{name}_{datetime.now().strftime('%H%M%S')}{ext}")
    shutil.move(pdf_path, dest_path)

def find_awb_pdfs() -> list:
    """Sab renamed AWB PDFs — mtime order (WhatsApp download order)."""
    files = [
        p for p in glob.glob(os.path.join(WATCH_FOLDER,"*.pdf"))
        if re.search(r"_\d{3}-\d{8}\.pdf$", os.path.basename(p), re.IGNORECASE)
    ]
    return sorted(files, key=lambda f: os.path.getmtime(f))

# ─── PHASE 3: DAILY FILL ──────────────────────────────────────────────────────

# ── PATCH 16: read_extract_rows — sheet_name param added ──────────────────────
def read_extract_rows(sheet_name: str = None) -> list:
    if sheet_name is None:
        sheet_name = EXTRACT_SHEET
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXTRACT_PATH, data_only=True)
        ws = wb[sheet_name]
        rows = []
        for r in range(2, ws.max_row+1):
            awb = ws.cell(row=r,column=1).value
            if not awb or str(awb).strip()=="": continue
            rd = {c: ws.cell(row=r,column=c).value for c in range(1,14)}  # A-M (13 cols)
            try: rd[12] = int(float(str(rd[12] or "0").replace(",","")))
            except (ValueError,TypeError): pass
            rows.append(rd)
        wb.close()
        return rows
    except Exception as e:
        print(f"[EXTRACT READ ERROR] {e}")
        return []

XL_UP = -4162

def _daily_last_row(ws_daily) -> int:
    return ws_daily.Cells(ws_daily.Rows.Count, DAILY_AWB_COL).End(XL_UP).Row

def _daily_next_empty(ws_daily) -> int:
    last = _daily_last_row(ws_daily)
    return last + 1

def _daily_existing_awbs(ws_daily) -> set:
    last = _daily_last_row(ws_daily)
    if last < 2:
        return set()
    rng = ws_daily.Range(
        ws_daily.Cells(2, DAILY_AWB_COL),
        ws_daily.Cells(last, DAILY_AWB_COL)
    ).Value
    if rng is None:
        return set()
    if isinstance(rng, (list, tuple)) and isinstance(rng[0], (list, tuple)):
        return {str(row[0]).strip() for row in rng if row[0]}
    return {str(rng).strip()} if rng else set()

# CCU fill — unchanged
def fill_daily_row(ws_daily, target_row: int, data: dict):
    for sc,dc in STEP1:
        ws_daily.Cells(target_row, dc).Value = data.get(sc,"")
    time.sleep(MACRO_WAIT)

    for sc,dc in STEP2:
        val = data.get(sc,"")
        if dc == 16:
            try: rate_val = float(str(val).replace(",","") or 0)
            except (ValueError,TypeError): rate_val = 0
            ws_daily.Cells(target_row,dc).Value = rate_val
            if rate_val >= FLAT_RATE_LIMIT:
                ws_daily.Cells(target_row, BASIC_COL_DAILY).Value = rate_val
                print(f"      Flat rate {rate_val:.0f} → Q = {rate_val:.0f}")
        elif dc == 19:
            try: ws_daily.Cells(target_row,dc).Value = int(float(str(val) or 0))
            except (ValueError,TypeError): ws_daily.Cells(target_row,dc).Value = val
        else:
            ws_daily.Cells(target_row,dc).Value = val

    for sc,dc in STEP3:
        val = data.get(sc,"")
        if dc == 3:
            try:
                from datetime import datetime as _dt
                d = _dt.strptime(str(val), "%d-%m-%Y")
                excel_serial = (d - _dt(1899, 12, 30)).days
                cell = ws_daily.Cells(target_row, dc)
                cell.Value        = excel_serial
                cell.NumberFormat = "DD-MM-YYYY"
                continue
            except (ValueError, TypeError):
                pass
        ws_daily.Cells(target_row,dc).Value = val

    for sc,dc in STEP_LAST:
        ws_daily.Cells(target_row,dc).Value = data.get(sc,"")

# ── PATCH 17: fill_daily_row_blr — nayi BLR fill function ─────────────────────
def fill_daily_row_blr(ws_daily, target_row: int, data: dict):
    """
    BLR Daily Sheet fill:
    Step 1 — AWB → B(2)   [macro trigger — AWB first]
    Step 2 — Agent→J, Material→K, PCS→L, CHW→M, Rate→N, DueAgent→P, DueCarrier→Q
    Step 3 — Date→C, Dest→E, Flight→F
    Last   — Consignor → I(9)
    """
    # Step 1 — AWB → B (macro trigger)
    for sc, dc in BLR_STEP1:
        ws_daily.Cells(target_row, dc).Value = data.get(sc, "")
    time.sleep(MACRO_WAIT)

    # Step 2 — Agent, Material, PCS, CHW, Rate, DueAgent, DueCarrier
    for sc, dc in BLR_STEP2:
        val = data.get(sc, "")
        if dc == 14:   # N = Rate
            try: rate_val = float(str(val).replace(",", "") or 0)
            except (ValueError, TypeError): rate_val = 0
            ws_daily.Cells(target_row, dc).Value = rate_val
            if rate_val >= FLAT_RATE_LIMIT:
                ws_daily.Cells(target_row, BLR_BASIC_COL_DAILY).Value = rate_val
                print(f"      BLR Flat rate {rate_val:.0f} → O = {rate_val:.0f}")
        elif dc == 17:  # Q = Due Carrier (int)
            try: ws_daily.Cells(target_row, dc).Value = int(float(str(val) or 0))
            except (ValueError, TypeError): ws_daily.Cells(target_row, dc).Value = val
        else:
            ws_daily.Cells(target_row, dc).Value = val

    # Step 3 — Date→C, Dest→E, Flight→F
    for sc, dc in BLR_STEP3:
        val = data.get(sc, "")
        if dc == 3:   # C = Date — Excel serial
            try:
                from datetime import datetime as _dt
                d = _dt.strptime(str(val), "%d-%m-%Y")
                excel_serial = (d - _dt(1899, 12, 30)).days
                cell = ws_daily.Cells(target_row, dc)
                cell.Value        = excel_serial
                cell.NumberFormat = "DD-MM-YYYY"
                continue
            except (ValueError, TypeError):
                pass
        ws_daily.Cells(target_row, dc).Value = val

    # Last — Consignor → I(9)
    for sc, dc in BLR_STEP_LAST:
        ws_daily.Cells(target_row, dc).Value = data.get(sc, "")

# ── PATCH 18: _do_extract_phase — sab PDFs ek saath, origin-wise route ────────
def _do_extract_phase() -> tuple:
    """
    Watch folder ke sab renamed PDFs:
    1. Extract karo
    2. Origin detect karo
    3. Sahi sheet (CCU/BLR) mein save karo
    4. Sahi PROCESSED folder mein move karo
    Returns: (ccu_added, blr_added)
    """
    pdfs = find_awb_pdfs()
    if not pdfs:
        print("\n[EXTRACT] Koi nayi PDF nahi.")
        return 0, 0

    print(f"\n[EXTRACT] {len(pdfs)} renamed PDF(s) mili...")
    ccu_added = blr_added = 0

    for pdf_path in pdfs:
        fname = os.path.basename(pdf_path)
        print(f"  ► {fname}")
        data = extract_from_pdf(pdf_path)
        if not data or not data.get("awb_no"):
            print(f"    ❌ Extract fail"); continue

        # PDF se origin detect karo
        text   = read_pdf_page1(pdf_path)
        origin = detect_origin(text)
        if origin not in ("CCU", "BLR"):
            origin = "CCU"   # default fallback
            print(f"    [ORIGIN] Unknown → CCU (default)")
        else:
            print(f"    [ORIGIN] {origin}")

        cfg    = get_origin_cfg(origin)
        result = save_to_extract(data, cfg["extract_sheet"])

        if result == "added":
            move_to_processed(pdf_path, cfg["processed_folder"])
            print(f"    ✅ [{origin}] Extract → {cfg['extract_sheet']} + {os.path.basename(cfg['processed_folder'])}")
            if origin == "BLR": blr_added += 1
            else: ccu_added += 1
        elif result == "duplicate":
            print(f"    ⏭  [{origin}] Already in extract")
        else:
            print(f"    ❌ [{origin}] Error")

    return ccu_added, blr_added

# ── PATCH 19: _do_fill_phase — ek origin ka daily sheet fill ──────────────────
def _do_fill_phase(origin: str, excel, wb_daily) -> tuple:
    """
    Specified origin ka daily sheet fill karo.
    Returns: (added_list, failed_list)
    """
    cfg        = get_origin_cfg(origin)
    sheet_name = cfg["daily_sheet_name"]
    fill_fn    = fill_daily_row_blr if origin == "BLR" else fill_daily_row

    extract_rows = read_extract_rows(cfg["extract_sheet"])
    if not extract_rows:
        print(f"  [{origin}] Extract sheet mein koi data nahi.")
        return [], []

    # Daily sheet find karo
    ws_daily = None
    for sh in wb_daily.Sheets:
        if sh.Name.strip().lower() == sheet_name.strip().lower():
            ws_daily = sh; break

    if ws_daily is None:
        names = [sh.Name for sh in wb_daily.Sheets]
        _popup("error", "Sheet Error",
               f"[{origin}] Sheet '{sheet_name}' nahi mili!\n"
               f"Mili sheets: {', '.join(names)}")
        return [], []

    ws_daily.Activate()
    existing = _daily_existing_awbs(ws_daily)
    new_rows = [r for r in extract_rows if str(r.get(1,"")).strip() not in existing]

    if not new_rows:
        print(f"  [{origin}] Koi naya AWB nahi.")
        return [], []

    print(f"  [{origin}] {len(new_rows)} naye AWBs fill honge...")
    added = []; failed = []

    for rd in new_rows:
        awb = str(rd.get(1,"")).strip()
        ag  = str(rd.get(6,"")).strip()
        print(f"  ► [{origin}] AWB:{awb}  Agent:{ag}")
        try:
            target = _daily_next_empty(ws_daily)
            fill_fn(ws_daily, target, rd)
            print(f"    ✅ Row {target}")
            added.append(awb)
            memory_add_awb(awb, ag)
        except Exception as e:
            print(f"    ❌ {e}")
            failed.append(awb)

    return added, failed

# ── PATCH 20: _get_or_open_excel — Excel open/reuse common logic ──────────────
def _get_or_open_excel():
    """Existing Excel instance reuse karo ya naya open karo."""
    import win32com.client
    daily_name = os.path.basename(DAILY_PATH).lower()
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        print("  [EXCEL] Existing instance mila.")
    except Exception:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        print("  [EXCEL] Naya instance khola.")

    excel.DisplayAlerts = False
    excel.EnableEvents  = True

    wb_daily = None
    for wb in excel.Workbooks:
        if os.path.basename(wb.FullName).lower() == daily_name:
            wb_daily = wb
            print("  [EXCEL] DAILYSHEET already open — reuse.")
            break
    if wb_daily is None:
        wb_daily = excel.Workbooks.Open(DAILY_PATH)
        print("  [EXCEL] DAILYSHEET naya open kiya.")

    return excel, wb_daily

# ── PATCH 21: run_daily_fill — origin param se CCU ya BLR fill ────────────────
def run_daily_fill(origin: str = "CCU"):
    """Phase 2 (extract all) + Phase 3 (fill specified origin)."""
    import win32com.client, pythoncom
    pythoncom.CoInitialize()

    # Phase 2 — sab PDFs extract karo (CCU + BLR dono automatically route honge)
    _do_extract_phase()

    # Phase 3 — sirf is origin ka daily sheet fill karo
    cfg = get_origin_cfg(origin)
    extract_rows = read_extract_rows(cfg["extract_sheet"])
    if not extract_rows:
        _popup("info", "Koi Data Nahi",
               f"[{origin}] Extract sheet mein koi data nahi mila.")
        return

    excel = wb_daily = None
    try:
        excel, wb_daily = _get_or_open_excel()

        added, failed = _do_fill_phase(origin, excel, wb_daily)

        if not added and not failed:
            _popup("info", "Kuch Nahi",
                   f"[{origin}] Koi naya AWB fill karne ke liye nahi mila.")
            wb_daily.Close(SaveChanges=False)
            return

        wb_daily.Save()
        print(f"\n[DONE] [{origin}] Added:{len(added)}  Failed:{len(failed)}")

        summary = (f"[{origin}] ✅  {len(added)} AWB(s) fill hue\n"
                   f"❌  {len(failed)} fail hue")
        if failed: summary += "\n\nFailed:\n" + "\n".join(failed)
        _popup("info", f"{origin} Fill Complete ✔", summary)

    except Exception as e:
        print(f"[{origin} DAILY ERROR] {e}")
        _popup("error", "Error", f"[{origin}] Daily fill error:\n{e}")
    finally:
        try:
            if excel: excel.DisplayAlerts = True
        except Exception: pass
        try: pythoncom.CoUninitialize()
        except Exception: pass

# ── PATCH 22: run_both_fill — CCU + BLR sequentially, ek hi Excel instance ────
def run_both_fill():
    """
    Phase 2 (extract all) + Phase 3 (CCU fill) + Phase 3 (BLR fill).
    Ek hi Excel instance — sequential, koi conflict nahi.
    """
    import win32com.client, pythoncom
    pythoncom.CoInitialize()

    # Phase 2 — ek baar extract sab
    ccu_ext, blr_ext = _do_extract_phase()
    print(f"\n[BOTH] Extract done — CCU:{ccu_ext}  BLR:{blr_ext}")

    excel = wb_daily = None
    try:
        excel, wb_daily = _get_or_open_excel()

        # CCU fill pehle
        print("\n[BOTH] CCU fill shuru...")
        ccu_added, ccu_failed = _do_fill_phase("CCU", excel, wb_daily)

        # BLR fill baad mein — same Excel, same workbook
        print("\n[BOTH] BLR fill shuru...")
        blr_added, blr_failed = _do_fill_phase("BLR", excel, wb_daily)

        wb_daily.Save()
        print(f"\n[BOTH DONE] CCU Added:{len(ccu_added)} Failed:{len(ccu_failed)} | "
              f"BLR Added:{len(blr_added)} Failed:{len(blr_failed)}")

        summary = (
            f"✅ CCU: {len(ccu_added)} fill hue  ❌ {len(ccu_failed)} fail\n"
            f"✅ BLR: {len(blr_added)} fill hue  ❌ {len(blr_failed)} fail"
        )
        if ccu_failed: summary += f"\n\nCCU Failed: {', '.join(ccu_failed)}"
        if blr_failed: summary += f"\n\nBLR Failed: {', '.join(blr_failed)}"
        _popup("info", "Fill Both Complete ✔", summary)

    except Exception as e:
        print(f"[BOTH FILL ERROR] {e}")
        _popup("error", "Error", f"Fill Both error:\n{e}")
    finally:
        try:
            if excel: excel.DisplayAlerts = True
        except Exception: pass
        try: pythoncom.CoUninitialize()
        except Exception: pass

# ── PATCH 23: process_folder_only — unchanged logic, extract se handle hoga ───
def process_folder_only():
    """
    Watch folder ke sab renamed PDFs extract karo.
    Origin automatically detect hokar route hoga.
    """
    pdfs = find_awb_pdfs()
    if not pdfs:
        _popup("info", "Koi File Nahi",
               f"Koi renamed AWB PDF nahi mila:\n{WATCH_FOLDER}\n\n"
               "File naam hona chahiye:\nAGENTNAME_312-XXXXXXXX.pdf")
        return

    mem = memory_load()
    last_info = ""
    if mem.get("last_awb"):
        last_info = (
            f"📌 Pichhla processed AWB:\n"
            f"   AWB   : {mem['last_awb']}\n"
            f"   Agent : {mem['last_agent']}\n"
            f"   Time  : {mem['last_time']}\n\n"
            f"{'─'*38}\n\n"
        )

    names = "\n".join(
        f"{i+1}. {os.path.basename(p)}"
        for i, p in enumerate(pdfs)
    )
    ans = _ask_yesno_main(
        f"Extract PDFs ({len(pdfs)} files)",
        f"{last_info}{len(pdfs)} renamed PDF(s) mili (download order):\n\n{names}\n\n"
        f"CCU/BLR automatically detect hokar extract honge. Process karein?"
    )
    if not ans:
        return

    import pythoncom
    pythoncom.CoInitialize()
    try:
        ccu_added, blr_added = _do_extract_phase()
        _popup("info", "Extract Complete ✔",
               f"✅ CCU: {ccu_added} PDFs extracted\n"
               f"✅ BLR: {blr_added} PDFs extracted\n\n"
               "Ab Fill CCU / Fill BLR / Fill Both se daily sheet fill karo.")
    finally:
        try: pythoncom.CoUninitialize()
        except Exception: pass

# ─── SYSTEM TRAY ──────────────────────────────────────────────────────────────

def make_tray_icon():
    from PIL import Image, ImageDraw
    img = Image.new("RGBA",(64,64),(0,0,0,0))
    d   = ImageDraw.Draw(img)
    d.ellipse([0,0,63,63], fill="#1C1F2E")
    d.polygon([(32,8),(20,40),(32,34),(44,40)], fill="#3D7EFF")
    d.polygon([(12,32),(32,26),(52,32),(32,36)], fill="#5C96FF")
    return img

# ── PATCH 24: build_tray — updated menu with CCU/BLR/Both ────────────────────
def build_tray():
    import pystray

    def on_fill_ccu(icon, item):
        print("\n[TRAY] Fill CCU Daily Sheet clicked")
        threading.Thread(target=run_daily_fill, args=("CCU",), daemon=True).start()

    def on_fill_blr(icon, item):
        print("\n[TRAY] Fill BLR Daily Sheet clicked")
        threading.Thread(target=run_daily_fill, args=("BLR",), daemon=True).start()

    def on_fill_both(icon, item):
        print("\n[TRAY] Fill Both clicked")
        threading.Thread(target=run_both_fill, daemon=True).start()

    def on_process(icon, item):
        print("\n[TRAY] Extract PDFs clicked")
        threading.Thread(target=process_folder_only, daemon=True).start()

    def on_last(icon, item):
        threading.Thread(target=memory_show_last, daemon=True).start()

    def on_reset(icon, item):
        threading.Thread(target=memory_clear_processed, daemon=True).start()

    def on_exit(icon, item):
        print("\n[TRAY] Exit clicked — saving memory...")
        icon.stop()
        os._exit(0)

    menu = pystray.Menu(
        pystray.MenuItem("✈ PCCS AWB System", None, enabled=False),
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("📋 Fill CCU Sheet",       on_fill_ccu),
        pystray.MenuItem("📋 Fill BLR Sheet",       on_fill_blr),
        pystray.MenuItem("📋 Fill Both (CCU + BLR)", on_fill_both),
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("📂 Extract PDFs Only",    on_process),
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("🕐 Last Session Info",    on_last),
        pystray.MenuItem("🗑 Reset Memory",         on_reset),
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("❌ Exit",                 on_exit),
    )
    icon = pystray.Icon("PCCS", make_tray_icon(), "PCCS AWB System", menu)
    global _tray_icon_ref
    _tray_icon_ref = icon   # _notify() ke liye globally accessible
    return icon

# ─── FIRST RUN SETUP WIZARD ───────────────────────────────────────────────────

def _browse_folder(var, root):
    from tkinter import filedialog
    path = filedialog.askdirectory(parent=root, title="Folder Select Karo")
    if path:
        var.set(path.replace("/", "\\"))

def _browse_file(var, root, filetypes):
    from tkinter import filedialog
    path = filedialog.askopenfilename(parent=root, filetypes=filetypes)
    if path:
        var.set(path.replace("/", "\\"))

def run_setup_wizard() -> bool:
    """
    Pehli baar chalane par paths set karne ka GUI form.
    Returns True agar setup complete hua, False agar cancel kiya.
    """
    BG    = "#0F1117"; CARD  = "#1C1F2E"; BORDER = "#2A2D45"
    ACC   = "#3D7EFF"; ACC_H = "#5C96FF"
    TEXT  = "#E8EBF5"; MUTED = "#737AA8"; EBG = "#141622"

    result = [False]

    root = tk.Tk()
    root.title("PCCS — First Time Setup")
    root.configure(bg=BG)
    root.resizable(False, False)
    root.attributes("-topmost", True)

    W, H = 600, 590
    sw = root.winfo_screenwidth(); sh = root.winfo_screenheight()
    root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")

    # ── Header ────────────────────────────────────────────────────────────────
    tk.Frame(root, bg=ACC, height=3).pack(fill="x")
    hdr = tk.Frame(root, bg=CARD, padx=24, pady=14)
    hdr.pack(fill="x")
    tk.Label(hdr, text="✈  PCCS AWB System — First Time Setup",
             font=("Segoe UI", 13, "bold"), bg=CARD, fg=TEXT).pack(anchor="w")
    tk.Label(hdr, text="Apne computer ke folders aur files ek baar select karo — phir dobara nahi puchega",
             font=("Segoe UI", 8), bg=CARD, fg=MUTED).pack(anchor="w")
    tk.Frame(root, bg=BORDER, height=1).pack(fill="x")

    # ── Scrollable form area ───────────────────────────────────────────────────
    canvas = tk.Canvas(root, bg=BG, highlightthickness=0)
    canvas.pack(fill="both", expand=True)
    form = tk.Frame(canvas, bg=BG, padx=24, pady=14)
    canvas.create_window((0, 0), window=form, anchor="nw")

    # Field definitions: (label, hint, kind, filetypes_or_None)
    fields_def = [
        ("📥  Watch Folder",
         "Jahan PDF download hoti hai  (e.g.  D:\\MY DOWNLOADS)",
         "folder", None),
        ("📊  AWB Extract Excel  (.xlsx)",
         "AWB_EXTRACT-DATA.xlsx ka poora path  (CCU + BLR sheets isme honge)",
         "file",   [("Excel Files", "*.xlsx")]),
        ("📂  CCU Processed Folder",
         "CCU PDFs yahan move hongi  (e.g.  D:\\AWB_PDF\\PROCESSED)",
         "folder", None),
        ("📂  BLR Processed Folder",
         "BLR PDFs yahan move hongi  (e.g.  D:\\AWB_PDF\\PROCESSED_BLR)",
         "folder", None),
        ("📋  Daily Sheet  (.xlsm)",
         "PRADEEP_DAILYSHEET.xlsm ka poora path",
         "file",   [("Excel Macro Files", "*.xlsm")]),
    ]

    cfg_keys = [
        "watch_folder",
        "extract_excel",
        "processed_folder",
        "blr_processed_folder",
        "daily_sheet",
    ]

    vars_ = []
    for label, hint, kind, ft in fields_def:
        lf = tk.Frame(form, bg=BG); lf.pack(fill="x", pady=(0, 12))

        tk.Label(lf, text=label, font=("Segoe UI", 9, "bold"),
                 bg=BG, fg=TEXT).pack(anchor="w")
        tk.Label(lf, text=hint, font=("Segoe UI", 8),
                 bg=BG, fg=MUTED).pack(anchor="w")

        row = tk.Frame(lf, bg=BG); row.pack(fill="x", pady=(4, 0))

        v = tk.StringVar()
        vars_.append(v)

        ef = tk.Frame(row, bg=ACC, padx=1, pady=1)
        ef.pack(side="left", fill="x", expand=True)
        tk.Entry(ef, textvariable=v, font=("Segoe UI", 9),
                 bg=EBG, fg=TEXT, insertbackground=ACC,
                 bd=0, highlightthickness=0).pack(
                     fill="x", ipady=7, padx=8)

        if kind == "folder":
            cmd = lambda v=v: _browse_folder(v, root)
            btn_lbl = "📁 Browse"
        else:
            cmd = lambda v=v, ft=ft: _browse_file(v, root, ft)
            btn_lbl = "📄 Browse"

        bx = tk.Button(row, text=btn_lbl,
                       font=("Segoe UI", 8, "bold"),
                       bg=CARD, fg=TEXT,
                       activebackground=ACC, activeforeground="white",
                       bd=0, padx=12, pady=7, cursor="hand2",
                       command=cmd)
        bx.pack(side="left", padx=(8, 0))
        bx.bind("<Enter>", lambda e, b=bx: b.config(bg=ACC))
        bx.bind("<Leave>", lambda e, b=bx: b.config(bg=CARD))

    # Update scroll region after widgets are drawn
    form.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    # ── Bottom buttons ────────────────────────────────────────────────────────
    tk.Frame(root, bg=BORDER, height=1).pack(fill="x")
    btn_row = tk.Frame(root, bg=CARD, padx=24, pady=12)
    btn_row.pack(fill="x")

    def _save():
        # Validate — koi field khali nahi hona chahiye
        for i, v in enumerate(vars_):
            if not v.get().strip():
                messagebox.showwarning(
                    "⚠ Field Khali Hai",
                    f"Sab fields fill karna zaroori hai!\n\n"
                    f"Yeh field khali hai:\n{fields_def[i][0]}",
                    parent=root)
                return

        # Config dict banao
        cfg_new = {
            "paths": {
                k: v.get().strip()
                for k, v in zip(cfg_keys, vars_)
            },
            "settings": {
                "extract_sheet"       : "CCU",
                "blr_extract_sheet"   : "BLR",
                "daily_sheet_name"    : "ccu_2025-26",
                "blr_sheet_name"      : "BLR-2025-26",
                "macro_wait_sec"      : 0.6,
                "flat_rate_threshold" : 1000
            }
        }

        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(cfg_new, f, indent=2, ensure_ascii=False)
            result[0] = True
            messagebox.showinfo(
                "Setup Complete ✔",
                "✅  Setup complete ho gaya!\n\n"
                "PCCS ab normal start hoga.\n"
                "System tray mein ✈ icon dikhega.",
                parent=root)
            root.destroy()
        except Exception as e:
            messagebox.showerror("Save Error",
                f"Config save nahi hua:\n{e}", parent=root)

    def _cancel():
        if messagebox.askyesno("Exit?",
            "Setup cancel karne par PCCS band ho jayega.\nConfirm?",
            parent=root):
            root.destroy()

    tk.Button(btn_row, text="Cancel",
              font=("Segoe UI", 9), bg=CARD, fg=MUTED,
              activebackground=CARD, activeforeground=TEXT,
              bd=0, padx=14, pady=8, cursor="hand2",
              command=_cancel).pack(side="right", padx=(8, 0))

    sb = tk.Button(btn_row, text="  ✔  Save & Start  ",
                   font=("Segoe UI", 10, "bold"),
                   bg=ACC, fg="white",
                   activebackground=ACC_H, activeforeground="white",
                   bd=0, padx=22, pady=8, cursor="hand2",
                   command=_save)
    sb.pack(side="right")
    sb.bind("<Enter>", lambda e: sb.config(bg=ACC_H))
    sb.bind("<Leave>", lambda e: sb.config(bg=ACC))

    root.mainloop()
    return result[0]


# ─── STARTUP MESSAGE ──────────────────────────────────────────────────────────

def show_startup_info():
    mem = memory_load()
    msg = "✈ PCCS AWB System shuru ho gaya!\n\n"
    msg += f"Watching: {WATCH_FOLDER}\n"
    msg += f"CCU Agents: {len(AGENT_LIST_CCU)}  |  BLR Agents: {len(AGENT_LIST_BLR)}\n\n"
    if mem.get("last_awb"):
        msg += (f"📌 Last Session:\n"
                f"  AWB   : {mem['last_awb']}\n"
                f"  Agent : {mem['last_agent']}\n"
                f"  Time  : {mem['last_time']}\n\n")
    msg += "Taskbar mein ✈ icon pe right-click karein."

    root = tk.Tk(); root.withdraw()
    root.attributes("-topmost",True)
    messagebox.showinfo("PCCS Started", msg, parent=root)
    root.destroy()

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    # ── STEP 1: First-run check — config nahi hai to wizard chalao ────────────
    if not os.path.isfile(CONFIG_FILE):
        print("[SETUP] config.json nahi mila — Setup Wizard khul raha hai...")
        ok = run_setup_wizard()
        if not ok:
            print("[EXIT] Setup cancel — band ho raha hoon.")
            sys.exit(0)

    # ── STEP 2: Globals load karo (wizard ke baad bhi sahi kaam karega) ───────
    try:
        _init_globals()
    except Exception as e:
        _fatal(f"Config load error:\n{e}\n\nconfig.json check karo ya delete\nkaro taaki Setup Wizard dobara chale.")

    # ── STEP 3: Path validation ────────────────────────────────────────────────
    errors = []
    if not os.path.isdir(WATCH_FOLDER):
        errors.append(f"Watch folder nahi mila:\n  {WATCH_FOLDER}")
    if not os.path.isfile(EXTRACT_PATH):
        errors.append(f"Extract Excel nahi mila:\n  {EXTRACT_PATH}")
    if not os.path.isfile(DAILY_PATH):
        errors.append(f"DAILYSHEET nahi mila:\n  {DAILY_PATH}")
    if errors:
        # Config delete karo taaki wizard dobara chale
        try: os.remove(CONFIG_FILE)
        except Exception: pass
        _fatal("Path Error — Setup dobara chalega:\n\n" +
               "\n\n".join(errors) +
               "\n\nApp dobara open karo.")

    print("╔══════════════════════════════════════════╗")
    print("║    PCCS AWB Automation System v2.0       ║")
    print("║    CCU + BLR Dual Origin Support         ║")
    print("╠══════════════════════════════════════════╣")
    print(f"║  Watch  : {WATCH_FOLDER[:32]:<32}║")
    print(f"║  CCU Agents : {len(AGENT_LIST_CCU):<28}║")
    print(f"║  BLR Agents : {len(AGENT_LIST_BLR):<28}║")
    print("╚══════════════════════════════════════════╝\n")

    threading.Thread(target=show_startup_info, daemon=True).start()

    observer = Observer()
    observer.schedule(PDFHandler(), WATCH_FOLDER, recursive=False)
    observer.start()
    print(f"[WATCHER] Started → {WATCH_FOLDER}")

    tray = build_tray()
    tray_thread = threading.Thread(target=tray.run, daemon=True)
    tray_thread.start()
    print("[TRAY] System tray icon active")

    try:
        while True:
            try:
                fn = _popup_queue.get(timeout=0.1)
                try:
                    fn()
                except Exception as e:
                    print(f"[POPUP ERROR] {e}")   # Loop crash nahi hoga
            except queue.Empty:
                pass
    except KeyboardInterrupt:
        pass
    finally:
        observer.stop()
        observer.join()
        print("\n[STOPPED] PCCS band ho gaya.")

if __name__ == "__main__":
    main()
